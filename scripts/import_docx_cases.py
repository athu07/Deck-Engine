# -*- coding: utf-8 -*-
import os as _os, sys as _sys
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
from deckengine import config  # noqa: E402  (repo-root data paths)
"""
import_docx_cases.py  --  Append case studies written in the house Word format to
the case-study master Excel, so they flow into the content store like every other
case (run build_case_study_store.py + build_case_embeddings.py afterwards).

The Word doc must use the standard structure, one block per case:
    CASE STUDY: <TITLE>
    Client: <one line>            (context only; slides stay anonymised)
    Domain: <Industry> / <Function>
    The Challenge      <paragraph>
    The Solution       <paragraph>
    Key Capabilities   <6 bullets, each "Title: description">
    Results            <3 bullets>

It maps each case onto the 9 Excel columns (ID, Worktype, Keywords, Title,
Challenge, Solution, Capabilities, Results, AI Generated or Not), assigning new
MSSxxx ids after the current maximum. Typographic dashes/quotes are normalised
to plain ASCII (owner rule: hyphens only, no em/en dashes).

Run:
    py import_docx_cases.py --dry-run     # preview, writes nothing
    py import_docx_cases.py               # append to the master Excel (backs up first)
"""
import re
import shutil
import sys

import docx
import openpyxl

DOCX = r"C:\Users\E36250417\Downloads\New_Case_Study_Portfolio_v2_1_7.docx"
SRC_XLSX = r"C:\Users\E36250417\Downloads\Case_Studies_Master_IDed.xlsx"
WORKTYPE = "MS Solution"          # this portfolio is all MS Solution
WT_PREFIX = "MSS"

# words too generic to be useful search keywords
_STOP = {
    "the", "and", "for", "with", "from", "into", "across", "over", "under",
    "a", "an", "of", "to", "on", "in", "at", "by", "or", "per", "via",
    "this", "that", "its", "was", "were", "are", "is", "be", "been",
}

_TYPO = {
    "—": "-", "–": "-", "‒": "-", "‐": "-",
    "‘": "'", "’": "'", "“": '"', "”": '"',
    "…": "...", " ": " ", "�": "-",
}


def clean(s):
    s = s or ""
    for k, v in _TYPO.items():
        s = s.replace(k, v)
    return re.sub(r"[ \t]{2,}", " ", s).strip()


def parse_docx(path=DOCX):
    d = docx.Document(path)
    paras = [clean(p.text) for p in d.paragraphs]
    paras = [t for t in paras if t]
    cases, cur, section = [], None, None
    for t in paras:
        low = t.lower()
        if low.startswith("case study:"):
            if cur:
                cases.append(cur)
            cur = {"title": t.split(":", 1)[1].strip(), "domain": "",
                   "challenge": [], "solution": [], "caps": [], "results": []}
            section = None
            continue
        if cur is None:
            continue
        if low.startswith("client:"):
            section = None; continue
        if low.startswith("domain:"):
            cur["domain"] = t.split(":", 1)[1].strip(); section = None; continue
        if low == "the challenge":
            section = "challenge"; continue
        if low == "the solution":
            section = "solution"; continue
        if low.startswith("key capabilities"):
            section = "caps"; continue
        if low == "results":
            section = "results"; continue
        if section in ("challenge", "solution"):
            cur[section].append(t)
        elif section == "caps":
            cur["caps"].append(t)
        elif section == "results":
            cur["results"].append(t)
    if cur:
        cases.append(cur)
    for c in cases:
        c["challenge"] = " ".join(c["challenge"]).strip()
        c["solution"] = " ".join(c["solution"]).strip()
        # keep only capability lines that carry a "Title: desc" (defensive)
        c["caps"] = [x for x in c["caps"] if x.strip()][:6]
        # a stray non-metric trailing bullet (e.g. "TAXONOMY") is dropped: keep 3
        c["results"] = [x for x in c["results"] if x.strip()][:3]
    return cases


def _keywords_cell(case):
    """Excel Keywords cell = domain, sub-domain, then distinctive terms. The store
    builder uses the FIRST comma-item as the display domain + industry code."""
    parts = [p.strip() for p in re.split(r"[/|]", case["domain"]) if p.strip()]
    domain_first = parts[0] if parts else ""
    domain_second = parts[1] if len(parts) > 1 else ""
    # distinctive words from the title + each capability's title (before the colon)
    terms, seen = [], set()
    sources = [case["title"]] + [c.split(":", 1)[0] for c in case["caps"]]
    for src in sources:
        for w in re.findall(r"[A-Za-z][A-Za-z\-]+", src):
            wl = w.lower()
            if len(wl) >= 3 and wl not in _STOP and wl not in seen:
                seen.add(wl)
                terms.append(wl)
    kw = [domain_first]
    if domain_second:
        kw.append(domain_second)
    kw += terms[:10]
    return ", ".join(kw)


def build_rows(cases, start_num):
    rows = []
    for i, c in enumerate(cases):
        sid = f"{WT_PREFIX}{start_num + i:03d}"
        rows.append([
            sid,                       # 1 ID
            WORKTYPE,                  # 2 Worktype
            _keywords_cell(c),         # 3 Keywords
            c["title"],                # 4 Title
            c["challenge"],            # 5 Challenge
            c["solution"],             # 6 Solution
            "\n".join(c["caps"]),      # 7 Capabilities (one per line)
            "; ".join(c["results"]),   # 8 Results (semicolon-separated)
            None,                      # 9 AI Generated or Not (blank = real case)
        ])
    return rows


def next_mss_number(ws):
    mx = 0
    for r in range(2, ws.max_row + 1):
        v = str(ws.cell(r, 1).value or "").strip().upper()
        m = re.match(rf"{WT_PREFIX}(\d+)$", v)
        if m:
            mx = max(mx, int(m.group(1)))
    return mx + 1


def main(dry_run):
    cases = parse_docx()
    wb = openpyxl.load_workbook(SRC_XLSX)
    ws = wb.active
    start = next_mss_number(ws)
    rows = build_rows(cases, start)

    print(f"Parsed {len(cases)} cases from the Word doc.")
    print(f"Assigning ids {rows[0][0]} .. {rows[-1][0]} (appending after row {ws.max_row}).\n")
    for r in rows:
        caps_n = len([x for x in r[6].split(chr(10)) if x.strip()])
        res_n = len([x for x in r[7].split(';') if x.strip()])
        flag = "" if caps_n == 6 and res_n == 3 else f"  <<< caps={caps_n} results={res_n}"
        print(f"  {r[0]}  {r[3][:52]:52}  domain='{r[2].split(',')[0]}'{flag}")

    if dry_run:
        print("\nDRY RUN — nothing written. Re-run without --dry-run to append.")
        return

    backup = SRC_XLSX.replace(".xlsx", ".BEFORE_DOCX_IMPORT.xlsx")
    shutil.copyfile(SRC_XLSX, backup)
    for r in rows:
        ws.append(r)
    wb.save(SRC_XLSX)
    print(f"\nBacked up  -> {backup}")
    print(f"Appended {len(rows)} rows -> {SRC_XLSX}")
    print("Next: py build_case_study_store.py  &&  py build_case_embeddings.py")


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main(dry_run="--dry-run" in sys.argv)

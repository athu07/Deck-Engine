# -*- coding: utf-8 -*-
"""
build_case_study_store.py  --  Turn Case_Studies_Master.xlsx into a powerful,
queryable content store for content-referenced case-study slides.

WHAT IT DOES (read top-to-bottom, nothing hidden):
  1. Reads every case study from the master Excel (160 rows).
  2. Assigns a stable, human-readable, collision-free ID per row:
        AI Pods            -> AIP001 ..   Workforce Solutions -> WFS001 ..
        MS Solution        -> MSS001 ..
     (The legacy master deck uses CS01..CS136 — these new IDs never clash.)
  3. Splits the 6 capabilities and 3 results into clean lists (and FLAGS any
     row that does not have exactly 6 caps / 3 results so nothing breaks the
     6-card / 3-stat template).
  4. Derives STRONG tags by reusing the live engine's own brains:
        - work_type   (AI_POD / WORKFORCE / MS)            engine codes
        - industry    (extended taxonomy — covers Retail, Media, PE, EdTech…)
        - function    (tagger.FUNCTION)
        - persona     (personas.tag_slide)
        - keywords    (curated + synonym-expanded + concepts found in the text)
  5. Carries the "AI Generated or Not" flag (column 9) so AI-written case
     studies can be marked in "Create AI Slide" for future reference.
  6. Writes:
        case_study_content_store.json   (the content store the engine reads)
        Case_Studies_Master_IDed.xlsx   (your Excel, now with the ID column filled)
     Both are NEW files — your original Case_Studies_Master.xlsx is untouched.

Run:  py build_case_study_store.py
"""

import json
import re
import shutil

import openpyxl

import synonyms
import personas
from tagger import FUNCTION, _score

# SOURCE = the owner's hand-curated, ID'd master (now carries per-capability
# descriptions in the form "Title: description", one per line). Read-only here;
# we regenerate the engine-side mirror + JSON below and never touch the source.
SRC_XLSX = r"C:\Users\E36250417\Downloads\Case_Studies_Master_IDed.xlsx"
OUT_JSON = "case_study_content_store.json"
OUT_XLSX = "Case_Studies_Master_IDed.xlsx"   # engine-side mirror (project folder)

# Excel work-type label  ->  engine work-type code + ID prefix
WT_MAP = {
    "AI Pods":             ("AI_POD",    "AIP"),
    "Workforce Solutions": ("WORKFORCE", "WFS"),
    "MS Solution":         ("MS",        "MSS"),
}

# ---------------------------------------------------------------------------
# Industry: extend the engine's 8 codes to cover the richer domains in this
# data. Key = lowercase needle found in the raw domain string; value = code.
# Order matters — most specific first.
# ---------------------------------------------------------------------------
INDUSTRY_FROM_DOMAIN = [
    ("healthcare",            "HEALTHCARE"),
    ("medtech",               "HEALTHCARE"),
    ("medical",               "HEALTHCARE"),
    ("pharmaceutical",        "PHARMA"),
    ("pharma",                "PHARMA"),
    ("life sciences",         "PHARMA"),
    ("private equity",        "PRIVATE_EQUITY"),
    ("finance operations",    "FINANCE_OPS"),
    ("banking",               "BFSI"),
    ("insurance",             "BFSI"),
    ("aviation",              "AVIATION"),
    ("automotive",            "AUTOMOTIVE"),
    ("telecom",               "TELECOM"),
    ("process manufacturing", "PROCESS_MFG"),
    ("manufacturing",         "MANUFACTURING"),
    ("industrial",            "MANUFACTURING"),
    ("construction",          "CONSTRUCTION"),
    ("energy",                "ENERGY"),
    ("logistics",             "LOGISTICS"),
    ("education",             "EDTECH"),
    ("retail",                "RETAIL"),
    ("media",                 "MEDIA"),
    ("professional services", "PROF_SERVICES"),
    ("software",              "TECH_IT"),
    ("technology",            "TECH_IT"),
]

# The CLIENT | DOMAIN subheading already shows the client descriptor + domain,
# so the main TITLE should carry only the capability/engagement name — not a
# repeated "— A LEADING X COMPANY" / "— US TELECOM" / "— EDTECH" tail.
# A trailing dash-segment is dropped ONLY if, after removing generic/descriptor/
# domain words, nothing meaningful is left (so real tails like "CHEQUE CLEARING
# AUTOMATION" or "ORACLE OIPA" are kept).
_DESC_WORDS = {
    "a", "an", "the", "leading", "global", "major", "multinational", "large",
    "top", "company", "companies", "corporation", "corp", "enterprise",
    "institution", "firm", "group", "leader", "operator", "utility", "retailer",
    "provider", "manufacturer", "conglomerate", "mnc", "gcc", "coe",
    "organization", "organisation", "us", "uk", "usa", "apac", "emea", "india",
    "indian", "japanese", "european", "american", "north", "player", "giant",
}
# domain abbreviations that mean the same thing as a domain label
_DOMAIN_ALIASES = {
    "edtech", "fintech", "medtech", "healthtech", "insurtech", "proptech",
    "regtech", "bfsi", "telecom", "telco", "ott", "cpg", "fmcg", "oem",
}
# words too generic to count as real title content on their own
_GENERIC_TOKENS = {
    "engineering", "solution", "solutions", "system", "systems", "capability",
    "capabilities", "services", "service", "platform", "transformation",
    "delivery", "operations", "sector", "industry", "business", "division",
    "co", "domain",
}


# Owner rule: never use an em/en dash in slide content — always a plain hyphen.
def _dash(s):
    return (s or "").replace("—", "-").replace("–", "-")


def _tok(s):
    return set(re.findall(r"[a-z0-9]+", (s or "").lower()))


def _is_descriptor_segment(seg, domain):
    """True if the segment is purely a client/domain descriptor (no real content)."""
    words = _tok(seg)
    if not words:
        return False
    leftover = words - _DESC_WORDS - _DOMAIN_ALIASES - _GENERIC_TOKENS - _tok(domain)
    return len(leftover) == 0


def _clean_title(raw, domain=""):
    """Return (clean_title, descriptor). Drops a trailing dash-segment that only
    repeats the client/domain, leaving the capability name as the heading."""
    t = (raw or "").strip()
    descriptor = ""
    # remove an embedded "A LEADING ... COMPANY/INSTITUTION/..." descriptor phrase
    # (stops at a dash so it never eats the capability name on the other side)
    desc_noun = (r"(?:COMPANY|COMPANIES|INSTITUTION|FIRM|ENTERPRISE|CORPORATION|"
                 r"GROUP|LEADER|OPERATOR|UTILITY|RETAILER|PROVIDER|MANUFACTURER|"
                 r"CONGLOMERATE|MNC|ORGANI[SZ]ATION)")
    m = re.search(r"\bA LEADING\b[^—–\-]*?\b" + desc_noun + r"\b", t, re.I)
    if m:
        descriptor = m.group(0).strip()
        t = (t[:m.start()] + " " + t[m.end():]).strip()
    parts = re.split(r"\s+[—–-]\s+", t)
    while len(parts) > 1 and _is_descriptor_segment(parts[-1], domain):
        descriptor = parts[-1].strip() or descriptor
        parts = parts[:-1]
    t = " - ".join(parts)
    # fallback: an embedded "A LEADING ... " descriptor that has no dash before it
    m = re.search(r"\bA LEADING\b.*$", t, re.I)
    if m and _is_descriptor_segment(m.group(0), domain):
        descriptor = descriptor or t[m.start():].strip()
        t = t[:m.start()].strip()
    t = re.sub(r"\s*[—–-]\s*$", "", t).strip()
    t = re.sub(r"^\s*[—–-]\s*", "", t).strip()
    t = re.sub(r"\s{2,}", " ", t)
    return t, descriptor


def _split_semis(val):
    return [p.strip() for p in (val or "").split(";") if p.strip()]


def _split_caps(val):
    """Parse the Capabilities cell into [{title, body}, ...].

    The owner's format is one capability per LINE as 'Title: description'.
    First colon splits title from body; a line with no colon is title-only."""
    out = []
    for line in (val or "").splitlines():
        line = line.strip()
        if not line:
            continue
        if ":" in line:
            t, b = line.split(":", 1)
            out.append({"title": t.strip(), "body": b.strip()})
        else:
            out.append({"title": line, "body": ""})
    return out


def _split_commas(val):
    return [p.strip() for p in (val or "").split(",") if p.strip()]


def _industry_code(domain):
    low = (domain or "").lower()
    for needle, code in INDUSTRY_FROM_DOMAIN:
        if needle in low:
            return code
    return None


def _strong_keywords(curated, fulltext):
    """Build the searchable keyword set: curated terms + their synonyms + any
    known concept whose synonym group is hit anywhere in the case-study text."""
    bag = set()
    for k in curated:
        kl = k.strip().lower()
        if kl:
            bag |= synonyms.expand(kl)
    # add concepts the synonym index recognises in the body text
    for term in synonyms.known_terms():
        if len(term) < 3:
            continue
        if synonyms.hits_in(term, fulltext):
            bag |= synonyms.expand(term)
    # drop pure noise
    bag = {b for b in bag if len(b) >= 2}
    return sorted(bag)


def build():
    wb = openpyxl.load_workbook(SRC_XLSX)
    ws = wb.active

    counters = {"AIP": 0, "WFS": 0, "MSS": 0}
    records = []
    warnings = []

    for row in range(2, ws.max_row + 1):
        wt_label = (ws.cell(row=row, column=2).value or "").strip()
        if wt_label not in WT_MAP:
            continue
        wt_code, prefix = WT_MAP[wt_label]

        # prefer the owner's existing ID (col 1); only mint one if it's blank
        existing_id = (str(ws.cell(row=row, column=1).value or "")).strip()
        if existing_id:
            sid = existing_id
        else:
            counters[prefix] += 1
            sid = f"{prefix}{counters[prefix]:03d}"

        # normalise em/en dashes to plain hyphens on every text field
        raw_kw   = _dash(ws.cell(row=row, column=3).value or "")
        raw_title = _dash(ws.cell(row=row, column=4).value or "")
        challenge = _dash(ws.cell(row=row, column=5).value or "").strip()
        solution  = _dash(ws.cell(row=row, column=6).value or "").strip()
        caps_raw  = _dash(ws.cell(row=row, column=7).value or "")
        res_raw   = _dash(ws.cell(row=row, column=8).value or "")
        ai_flag_v = ws.cell(row=row, column=9).value

        kw_list = _split_commas(raw_kw)
        domain  = kw_list[0] if kw_list else ""
        curated = kw_list[1:] if len(kw_list) > 1 else []

        title, client_desc = _clean_title(raw_title, domain)
        caps = _split_caps(caps_raw)                  # [{title, body}, ...]
        caps_text = " ".join(c["title"] + " " + c["body"] for c in caps)
        results = _split_semis(res_raw)

        if len(caps) != 6:
            warnings.append(f"{sid}: {len(caps)} capabilities (template expects 6)")
        if len(results) != 3:
            warnings.append(f"{sid}: {len(results)} results (template expects 3)")

        industry = _industry_code(domain)

        # full text for deriving function / personas / concept keywords
        fulltext = " ".join([
            title, domain, " ".join(curated), challenge, solution,
            caps_text, " ".join(results),
        ]).lower()
        fulltext_padded = " " + fulltext + " "

        function, _ = _score(fulltext_padded, FUNCTION)

        persona_codes = personas.tag_slide({
            "primary_function": function or "",
            "work_types": wt_code,
            "keywords": " · ".join(curated),
            "title": title,
        })

        strong_kw = _strong_keywords(curated, fulltext)

        ai_generated = bool(str(ai_flag_v).strip().lower() in
                            ("y", "yes", "true", "ai", "ai generated", "1")) \
            if ai_flag_v is not None else False

        records.append({
            "id": sid,
            "work_type": wt_code,
            "work_type_label": wt_label,
            "title": title,
            "raw_title": raw_title.strip(),
            "client_descriptor": client_desc,   # anonymised; salesperson overrides
            "domain": domain,                    # raw industry label (for display)
            "industry": industry,                # engine code
            "function": function,
            "personas": persona_codes,
            "curated_keywords": curated,
            "keywords": strong_kw,               # synonym-expanded search fuel
            "challenge": challenge,
            "solution": solution,
            "capabilities": caps,                # expect 6
            "results": results,                  # expect 3
            "ai_generated": ai_generated,
            "source_row": row,
        })

        # write the ID back into the in-memory workbook (column 1)
        ws.cell(row=row, column=1).value = sid

    # save outputs (originals untouched — these are new files)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    wb.save(OUT_XLSX)

    # ---- report -----------------------------------------------------------
    print(f"Built {len(records)} case studies -> {OUT_JSON}")
    print(f"ID-stamped Excel        -> {OUT_XLSX}")
    by_wt = {}
    for r in records:
        by_wt[r["work_type_label"]] = by_wt.get(r["work_type_label"], 0) + 1
    print("\nBy work type:")
    for k, v in by_wt.items():
        print(f"  {k:22} {v}")

    miss_ind = [r["id"] for r in records if not r["industry"]]
    miss_fn  = [r["id"] for r in records if not r["function"]]
    print(f"\nIndustry derived : {len(records)-len(miss_ind)}/{len(records)}"
          + (f"  (no match: {miss_ind})" if miss_ind else ""))
    print(f"Function derived : {len(records)-len(miss_fn)}/{len(records)}"
          + (f"  (no match: {miss_fn})" if miss_fn else ""))

    avg_kw = sum(len(r["keywords"]) for r in records) / len(records)
    print(f"Avg search keywords/case: {avg_kw:.1f}")

    if warnings:
        print(f"\nStructure warnings ({len(warnings)}):")
        for w in warnings:
            print("  " + w)
    else:
        print("\nAll rows: clean 6 capabilities + 3 results.")


if __name__ == "__main__":
    build()

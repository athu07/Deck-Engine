# -*- coding: utf-8 -*-
"""
sync_deck_to_excel.py  --  Keep the case-study Excel in lock-step with the
LIVING working-copy deck.

Reads WORKING_COPY_Master_Deck.pptx (the deck the engine actually builds from),
pulls every case-study slide, and checks each one against the master Excel
(Case_Studies_Master.xlsx). Any case study that is ON THE DECK but NOT in the
Excel is reported (and, with --write, appended) so the Excel stays complete.

Re-runnable by design: run it again whenever a new slide is added to the deck
and only the genuinely-new cases get added — existing ones are matched and
skipped.

Usage:
  py sync_deck_to_excel.py            # REPORT ONLY (default, writes nothing)
  py sync_deck_to_excel.py --write    # also append missing cases to the Excel
"""

import re
import sys

import openpyxl

import build_library as bl

DECK = "WORKING_COPY_Master_Deck.pptx"
XLSX = "Case_Studies_Master_IDed.xlsx"      # the source of truth we keep complete
XLSX_OUT = "Case_Studies_Master_IDed.xlsx"  # overwrite in place when --write

MIDDOT = "·"

# Generic words that carry no identity — ignored when comparing titles.
_STOP = {
    "for", "the", "and", "with", "to", "of", "a", "an", "in", "on", "by", "at",
    "global", "leading", "client", "company", "major", "us", "from", "into",
    "case", "study", "engagement",
}


def _norm_tokens(text):
    """Significant lowercase tokens of a title (drops punctuation + stopwords)."""
    text = (text or "").lower()
    text = re.sub(r"[^a-z0-9 ]", " ", text)
    toks = [t for t in text.split() if len(t) >= 3 and t not in _STOP]
    return set(toks)


def _jaccard(a, b):
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


# ── deck-side parsing ──────────────────────────────────────────────────────

def _is_case_study(full_text):
    """A slide is a case study if it has the curated Challenge/Solution layout,
    is a generated gap slide, or carries a middot keyword subtitle."""
    if not full_text:
        return False
    if "THE CHALLENGE" in full_text or "Generated slide" in full_text:
        return True
    # middot keyword subtitle on the 2nd line
    lines = [l for l in full_text.splitlines() if l.strip()]
    return len(lines) >= 2 and lines[1].count(MIDDOT) >= 2


def _section(full_text, head, nexts):
    """Grab text between a section header and the next header (curated layout)."""
    if head not in full_text:
        return ""
    after = full_text.split(head, 1)[1]
    cut = len(after)
    for n in nexts:
        i = after.find(n)
        if i != -1:
            cut = min(cut, i)
    return after[:cut].strip()


def parse_deck_cases(path):
    recs = bl.build(path)
    cases = []
    for r in recs:
        if not _is_case_study(r["full_text"]):
            continue
        ft = r["full_text"]
        generated = "Generated slide" in ft
        challenge = _section(ft, "THE CHALLENGE",
                              ["THE SOLUTION", "KEY CAPABILITIES"])
        solution = _section(ft, "THE SOLUTION", ["KEY CAPABILITIES", "⚙"])
        cases.append({
            "slide_id": r["slide_id"],
            "title": r["title"],
            "subtitle": r["subtitle"],
            "full_text": ft,
            "generated": generated,
            "challenge": challenge,
            "solution": solution,
            "tokens": _norm_tokens(r["title"]),
        })
    return cases


# ── excel-side ─────────────────────────────────────────────────────────────

def load_excel_titles(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in range(2, ws.max_row + 1):
        wt = (ws.cell(row=row, column=2).value or "").strip()
        title = (ws.cell(row=row, column=4).value or "").strip()
        if not wt and not title:
            continue
        rows.append({"row": row, "wt": wt, "title": title,
                     "tokens": _norm_tokens(title)})
    return wb, ws, rows


def best_match(deck_case, excel_rows):
    best, score = None, 0.0
    for er in excel_rows:
        s = _jaccard(deck_case["tokens"], er["tokens"])
        if s > score:
            best, score = er, s
    return best, score


def main():
    write = "--write" in sys.argv
    MATCH_THRESHOLD = 0.45      # below this = treated as not in the Excel

    deck_cases = parse_deck_cases(DECK)
    wb, ws, excel_rows = load_excel_titles(XLSX)

    matched, missing, weak = [], [], []
    for dc in deck_cases:
        er, score = best_match(dc, excel_rows)
        if score >= 0.7:
            matched.append((dc, er, score))
        elif score >= MATCH_THRESHOLD:
            weak.append((dc, er, score))      # probable match, worth eyeballing
        else:
            missing.append((dc, er, score))

    print(f"Deck case-study slides : {len(deck_cases)}")
    print(f"Excel case studies     : {len(excel_rows)}")
    print(f"  strong match (>=0.70) : {len(matched)}")
    print(f"  weak match (0.45-0.70): {len(weak)}")
    print(f"  MISSING (<0.45)       : {len(missing)}")

    if weak:
        print("\n--- WEAK matches (likely already in Excel, anonymised name) ---")
        for dc, er, s in weak:
            tag = " [GENERATED]" if dc["generated"] else ""
            print(f"  {dc['slide_id']:6} {s:.2f}  {dc['title'][:48]:48}{tag}")
            print(f"         -> excel row {er['row']}: {er['title'][:60]}")

    if missing:
        print("\n--- MISSING from Excel (on deck, no Excel match) ---")
        for dc, er, s in missing:
            tag = " [GENERATED PLACEHOLDER]" if dc["generated"] else ""
            print(f"  {dc['slide_id']:6} {s:.2f}  {dc['title']}{tag}")
            if er:
                print(f"         closest excel: {er['title'][:60]} ({s:.2f})")

    if write and missing:
        real_missing = [dc for dc, er, s in missing if not dc["generated"]]
        skipped_gen = [dc for dc, er, s in missing if dc["generated"]]
        wt_label_from_deck = {
            "AI_POD": "AI Pods", "WORKFORCE": "Workforce Solutions", "MS": "MS Solution",
        }
        appended = 0
        for dc in real_missing:
            r = ws.max_row + 1
            ws.cell(row=r, column=2).value = ""   # worktype left blank for review
            ws.cell(row=r, column=3).value = dc["subtitle"]
            ws.cell(row=r, column=4).value = dc["title"]
            ws.cell(row=r, column=5).value = dc["challenge"]
            ws.cell(row=r, column=6).value = dc["solution"]
            ws.cell(row=r, column=9).value = "Source: deck " + (dc["slide_id"] or "?")
            appended += 1
        wb.save(XLSX_OUT)
        print(f"\nAppended {appended} real missing case(s) to {XLSX_OUT}.")
        if skipped_gen:
            print(f"Skipped {len(skipped_gen)} GENERATED placeholder slide(s) "
                  f"(not real case studies): "
                  f"{[d['slide_id'] for d in skipped_gen]}")
    elif missing:
        print("\n(Report only. Re-run with --write to append the real missing cases.)")


if __name__ == "__main__":
    main()
